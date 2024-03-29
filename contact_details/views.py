from django.shortcuts import render
from django.contrib.auth import logout
from .models import *
from django.contrib.auth.models import User
from django.db.models import Q
from django.http import HttpResponseRedirect,HttpResponse
import pandas as pd
import openpyxl
from io import BytesIO
import re

##########GLOBAL VARS#########

role_choices = {
    "admin": "admin",
    "department head": "dept_head",
    "staff" : "staff",
    "trustee": "trustee",
}

email_regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
##########GLOBAL VARS END##########


##########HELPER FUNCTIONS START#################
def validate_data_point(data):
    #If not valid return False,"error message"

    if str(data[1])[-10:].isnumeric() != True:
        print("return")
        return False, "Invalid phone number"
    # if not re.fullmatch(email_regex, data[2]):
    #     return False, "Invalid email address"
    if not data[3].lower().strip() in role_choices:
        return False, "Invalid role - '" + data[3] + "'. Valid options are " + str(list(role_choices.keys()))
    print(data[5])
    if len(data) > 5 and data[5]!='None' and data[5]!='' and data[5]!=None and not trust.objects.filter(name=data[5].lower().strip()).exists():
        return False, "Trust - '" + data[5] + "' is invalid"
    if not department.objects.filter(name=data[4].lower().strip()).exists():
        return False, "Department - '" + data[4] + "' is invalid"
    return True,"no error"

def sanitize_phno(phno):
    return str(phno)[-10:]

#############HELPER FUNCTIONS END###################



# Create your views here.

def contact_list(request):
    if request.user.is_authenticated:
        user = user_detail.objects.filter(email=request.user.email).first()
        if user != None:
            relation_rule = relation_table.objects.all().values()[0]#TODO:Add failsafes everywhere
            print(relation_rule)
            if relation_rule[user.role] == 'self':
                user_details = [user]
            elif relation_rule[user.role] == 'dept_staff':
                user_details = user_detail.objects.filter(role='staff', department=user.department)
            elif relation_rule[user.role] == 'dept_trust':
                user_details = user_detail.objects.filter(role__in=['dept_head', 'staff'], department__in=department.objects.filter(trust=user.department.trust))
            elif relation_rule[user.role] == 'loc_depts':
                user_details = user_detail.objects.filter(location=user.department.location)
            elif relation_rule[user.role] == 'all':
                user_details = user_detail.objects.all()
        else:
            user_details = []
        all_depts = department.objects.all()
        data = {
            'user_details': user_details,
            'curr_user': user,
            'all_depts': all_depts,
        }
        return render(request, 'contact_list.html', context=data)
    else:
        return HttpResponseRedirect('/oauth/login/google-oauth2/?next=/')

def upload_bulk_contacts(request):
    if request.user.is_authenticated:
        if request.method != "POST" or not "fileInput" in request.FILES:
            data = {
                "status":0
            }
        else:
            wb = openpyxl.load_workbook(request.FILES["fileInput"])
            print(wb.sheetnames)
            error_list = []
            success_list = []
            for sheet in wb.sheetnames:
                worksheet = wb[sheet]
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    print(row_data)
                    row_data[1] = row_data[1].split('.')[0]
                    is_valid, reason = validate_data_point(row_data)
                    if not is_valid:
                        error_list.append({'data': row_data, 'status': reason, "row_idx": str(row[0].row)})
                        continue
                    elif user_detail.objects.filter(phno = sanitize_phno(row_data[1].split('.')[0])).exists():
                        error_list.append({'data': row_data, 'status': "Entry with same phone number already exists", "row_idx": str(row[0].row)})
                    else:
                        new_user_detail = user_detail(name=row_data[0],phno=sanitize_phno(row_data[1].split('.')[0]),email=row_data[2])
                        dept = department.objects.get(name=row_data[4].lower().strip())
                        new_user_detail.department = dept
                        new_user_detail.location = dept.location
                        new_user_detail.role =  role_choices[row_data[3].lower().strip()]
                        new_user_detail.save()
                        success_list.append({'data': row_data, 'status': "Successfully added", "row_idx": str(row[0].row)})
            data = {
                "status": 1,
                "error_list": error_list,
                "success_list": success_list
            }
            print(data)
        return render(request, 'contacts_upload_status.html', context=data)
    else:
        return HttpResponseRedirect('/oauth/login/google-oauth2/?next=/')

def upload_single_contact(request):
    if request.user.is_authenticated:
        if request.method != "POST":
            data = {
                "status":0
            }
        else:
            error_list = []
            success_list = []
            dept = user_detail.objects.filter(email=request.user.email).first().department
            row_data = [request.POST['name'],request.POST['phno'],request.POST['email'],request.POST['role'],dept.name]  #Needed in this format to multiplex using validate function
            is_valid, reason = validate_data_point(row_data)
            if not is_valid:
                error_list.append({'data': row_data, 'status': reason, "row_idx": str(1)})
            elif user_detail.objects.filter(phno = sanitize_phno(request.POST['phno'].split('.')[0])).exists():
                error_list.append({'data': row_data, 'status': "Entry with same phone number already exists", "row_idx": str(1)})
            else:
                new_user_detail = user_detail(name=request.POST['name'],phno=sanitize_phno(request.POST['phno']),email=request.POST['email'])
                new_user_detail.department = dept
                new_user_detail.role =  role_choices[request.POST['role'].lower().strip()]
                new_user_detail.save()
                success_list.append({'data': row_data, 'status': "Successfully added", "row_idx": str(1)})
            data = {
                "status": 1,
                "error_list": error_list,
                "success_list": success_list
            }
        return render(request, 'contacts_upload_status.html', context=data)
    else:
        return HttpResponseRedirect('/oauth/login/google-oauth2/?next=/')       

def keep_awake(request):
    return render(request, 'header.html')

def auth_logout(request):
    logout(request)

    return HttpResponseRedirect("/")